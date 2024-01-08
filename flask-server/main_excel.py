import nwc
import dcf
import yfinance
import wacc
import freecashflow
import fixed_assets
import statementFunct
import styleModule
import numpy as np
import requests
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE, FORMAT_PERCENTAGE
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import MergedCell
from string import ascii_uppercase




#Main Code
# Prompt the user for a ticker symbol


ticker_symbol = input("Please enter the ticker symbol: ")
start_year = input("Enter the start fiscal year (e.g., 2017): ")
end_year = input("Enter the end fiscal year (e.g., 2022): ")


# Convert start_year and end_year to integers
start_year = int(start_year)
end_year = int(end_year)


# Generate a list of years from start_year to end_year
years = list(range(start_year, end_year + 1))
years_str = ','.join(map(str, years)) # Convert the list of years to a comma-separated string




# Construct the URL with the user's input
url = f"https://backend.simfin.com/api/v3/companies/statements/compact?ticker={ticker_symbol}&statements=PL,BS,CF&period=FY&fyear={years_str}"


#url = "https://backend.simfin.com/api/v3/companies/statements/compact?ticker=AMZN&statements=PL,BS,CF&period=FY&fyear=2017%2C2018%2C2019%2C2020%2C2021%2C2022"


headers = {
"accept": "application/json",
"Authorization": "b1c124d9-f078-4887-abb9-d3504b54b23b"
}


response = requests.get(url, headers=headers)
jsonStr = json.loads(response.text)


data = jsonStr[0]['statements']
arrDataFrames = np.empty(len(data), dtype=object)
arrNames = np.empty(len(data), dtype=object)


for i in range(len(data)):
   arrNames[i] = statementFunct.getStatementName(data[i]['statement'])
   columns = data[i]['columns']
   # Extract data
   table_data = data[i]['data']
   neededColumns = statementFunct.getColumns(data[i]['statement'])
   # Create a DataFrame
   arrDataFrames[i] = statementFunct.process_statement(table_data, columns, neededColumns)


fixed_assets_Sheet = fixed_assets.AddSheet(start_year, end_year)
free_cash_flow_sheet = freecashflow.AddFreeCashFlowSheet(start_year, end_year)
nwc_sheet = nwc.AddNetWorkingCapitalSheet(start_year, end_year)
wacc_df = wacc.get_wacc_dataframe(ticker_symbol, start_year, end_year)
dcf_sheet = dcf.add_dcf_sheet(start_year, end_year, ticker_symbol)




# Write the filtered data to an Excel file with formatting
with pd.ExcelWriter(ticker_symbol + '_financial_statements.xlsx', engine='openpyxl') as writer:
# Register the named styles with the writer's workbook
   writer.book.add_named_style(styleModule.currency_style)
   writer.book.add_named_style(styleModule.fiscal_year_style)
   for i in range(len(data)):
   # Write the DataFrame to the worksheet
       arrDataFrames[i].to_excel(writer, sheet_name=arrNames[i], startrow=1)


       # Get the worksheet object
       worksheet = writer.sheets[arrNames[i]]
       # Apply bold formatting to specific rows
       statementFunct.apply_bold_to_specific_rows(worksheet, arrNames[i])
       styleModule.SetFinStyle(worksheet)
   # Auto-adjust column widths for each sheet
   for sheet_name in writer.book.sheetnames:
       statementFunct.auto_adjust_column_width(writer.book[sheet_name])














   #Fixed Assets Sheet
   fixed_assets_Sheet.to_excel(writer, sheet_name='Fixed Assets', startrow=1)
   fa_worksheet = writer.sheets['Fixed Assets']
   styleModule.SetOtherStyle(fa_worksheet)
   fixed_assets.auto_adjust_column_width(fa_worksheet)
   styleModule.unbold_category_names(fa_worksheet)
   for row in fa_worksheet.iter_rows(min_row=2, max_row=fa_worksheet.max_row):
       if row[0].value == "Ending PP&E":
           for cell in row:
               cell.font = Font(bold=True)
          
   styleModule.add_border_under_fiscal_year_row(fa_worksheet, 2)


   last_col = fa_worksheet.max_column


   # Set and style the "Fixed Asset Scedule" title cell
   title_row = 1
   title_cell = fa_worksheet.cell(row=title_row, column=1)
   title_cell.value = "Fixed Assets Schedule"
   title_cell.font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_cell.alignment = Alignment(horizontal='left', vertical='center')
      
   # Apply the fill color to all cells in the title row
   for col in range(1, last_col + 1):
       cell = fa_worksheet.cell(row=title_row, column=col)
       cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")


   styleModule.FApopulate_and_style_fiscal_years(fa_worksheet, start_year, end_year)


   styleModule.add_border_under_fiscal_year_row(fa_worksheet, 9)
  
   # Set and style the "Assumptions" title cell
   title_row = 8
   title_cell = fa_worksheet.cell(row=title_row, column=1)
   title_cell.value = "Assumptions"
   title_cell.font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_cell.alignment = Alignment(horizontal='left', vertical='center')


  
   # Apply the fill color to all cells in the title row
   for col in range(1, last_col + 1):
       cell = fa_worksheet.cell(row=title_row, column=col)
       cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")
      
   fixed_assets.apply_percentage_format_to_rows(fa_worksheet)
   end_column = fa_worksheet.max_column


   styleModule.set_data_cells_background(fa_worksheet, 2, 6, 1, end_column, "F8F7F7")
   styleModule.set_data_cells_background(fa_worksheet, 9, 11, 1, end_column, "ebebeb")
   fixed_assets.style_ending_ppe_row(fa_worksheet, 'F8F7F7')
  
  
  
  
  
  
   # Free Cash Flow Sheet




   fiscal_years = list(range(start_year + 1, end_year + 1)) + [f"{year}E" for year in range(end_year + 1, end_year + 6)]
   fiscal_years = list(map(str, fiscal_years))
  
  
   free_cash_flow_sheet.to_excel(writer, sheet_name='Free Cash Flow', startrow=1)
   fcf_worksheet = writer.sheets['Free Cash Flow']
   freecashflow.auto_adjust_column_width(fcf_worksheet)


  
   header_row_num = 2
   for col_num in range(3, len(fiscal_years) + 2):  # Adjust the range based on your fiscal year columns
       cell = fcf_worksheet.cell(row=header_row_num, column=col_num)
       cell.style = styleModule.fiscal_year_style


   # Apply the same style as other sheets to the Free Cash Flow sheet
   styleModule.SetOtherStyle(fcf_worksheet)  # Assuming this is the function for applying general styles


   styleModule.add_border_under_fiscal_year_row(fcf_worksheet, 2)
   freecashflow.apply_bold_to_specific_rows(fcf_worksheet, 'Free Cash Flow')


   last_col = fcf_worksheet.max_column


   title_row = 1
   title_cell = fcf_worksheet.cell(row=title_row, column=1)
   title_cell.value = "Unlevered Free Cash Flow (mm)"
   title_cell.font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_cell.alignment = Alignment(horizontal='center', vertical='center')
      
   # Apply the fill color to all cells in the title row
   for col in range(1, last_col + 1):
       cell = fcf_worksheet.cell(row=title_row, column=col)
       cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")


   freecashflow.apply_percentage_format_to_rows(fcf_worksheet)
   styleModule.populate_and_style_fiscal_years(fcf_worksheet, start_year, end_year)


   styleModule.add_border_under_fiscal_year_row(fcf_worksheet, 26)




   # Set and style the "Assumptions" title cell
   title_row = 25
   title_cell = fcf_worksheet.cell(row=title_row, column=1)
   title_cell.value = "Assumptions"
   title_cell.font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_cell.alignment = Alignment(horizontal='left', vertical='center')


   # Apply the fill color to all cells in the title row
   for col in range(1, last_col + 1):
       cell = fcf_worksheet.cell(row=title_row, column=col)
       cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")




   # Set background color for data cells in rows 1-22 across all data columns
   end_column = fcf_worksheet.max_column
   styleModule.set_data_cells_background(fcf_worksheet, 2, 22, 1, end_column, "F8F7F7")
   freecashflow.style_unlevered_free_cash_flow_row(fcf_worksheet, 'fdfd96')


   styleModule.set_data_cells_background(fcf_worksheet, 26, 34, 1, end_column, "ebebeb")










   # Net Working Capital sheet


   nwc_sheet.to_excel(writer, sheet_name='Net Working Capital', startrow=1)
   nwc_worksheet = writer.sheets['Net Working Capital']
   end_column = nwc_worksheet.max_column


   styleModule.SetOtherStyle(nwc_worksheet)


   freecashflow.auto_adjust_column_width(nwc_worksheet)




   styleModule.NWCpopulate_and_style_fiscal_years(nwc_worksheet, start_year, end_year)
   styleModule.add_border_under_fiscal_year_row(nwc_worksheet, 2)
   styleModule.add_border_under_fiscal_year_row(nwc_worksheet, 18)
   nwc.style_asset_row(nwc_worksheet, 'F8F7F7')
   nwc.style_liabilities_row(nwc_worksheet, 'F8F7F7')


   styleModule.set_data_cells_background(nwc_worksheet, 2, 13, 1, end_column, "F8F7F7")
   styleModule.set_data_cells_background(nwc_worksheet, 18, 30, 1, end_column, "ebebeb")
  
   nwc.apply_bold_to_specific_rows(nwc_worksheet, 'Net Working Capital')
  
  
   last_col = nwc_worksheet.max_column


   title_row = 1
   title_cell = nwc_worksheet.cell(row=title_row, column=1)
   title_cell.value = "Net Working Capital"
   title_cell.font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_cell.alignment = Alignment(horizontal='left', vertical='center')
      
   for col in range(1, last_col + 1):
       cell = nwc_worksheet.cell(row=title_row, column=col)
       cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")


   title_row = 17
   title_cell = nwc_worksheet.cell(row=title_row, column=1)
   title_cell.value = "Assumptions"
   title_cell.font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_cell.alignment = Alignment(horizontal='left', vertical='center')
      
   for col in range(1, last_col + 1):
       cell = nwc_worksheet.cell(row=title_row, column=col)
       cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")
   nwc.apply_percentage_format_to_rows(nwc_worksheet)










   #WACC sheet


   wacc_df.to_excel(writer, sheet_name='WACC', index=True)
   wacc_worksheet = writer.sheets['WACC']
      
   title_row = 1
   start_column = 1  # Column A
   end_column = 2    # Column B


   # Merge cells A1 and B1
   wacc_worksheet.merge_cells(start_row=title_row, start_column=start_column, end_row=title_row, end_column=end_column)




      
   wacc.auto_adjust_column_widths(wacc_worksheet)
   wacc.apply_percentage_format_to_rows(wacc_worksheet)
   wacc.wacc_Currency(wacc_worksheet)
   wacc.remove_borders_and_gridlines(wacc_worksheet)
   styleModule.set_data_cells_background(wacc_worksheet, 2, 13, 1, end_column, "F8F7F7")
   wacc.format_rows(wacc_worksheet)
   # Set the value for the merged cell
   title_cell = wacc_worksheet.cell(row=title_row, column=start_column)
   title_cell.value = "Weighted Average Cost of Capital"


   # Define the font, fill, and alignment
   title_font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_alignment = Alignment(horizontal='center', vertical='center')
   # Apply the style to the entire merged cell
   for col in range(start_column, end_column + 1):
       cell = wacc_worksheet.cell(row=title_row, column=col)
       cell.font = title_font
       cell.fill = title_fill
       cell.alignment = title_alignment
      


   #DCF
  
   dcf_sheet.to_excel(writer, sheet_name='DCF', startrow=1)
   dcf_worksheet = writer.sheets['DCF']
   freecashflow.auto_adjust_column_width(dcf_worksheet)
   styleModule.SetDCFStyle(dcf_worksheet)






   title_row = 2
   title_cell = dcf_worksheet.cell(row=title_row, column=1)
   title_cell.value = "Fiscal Year"
   title_cell.alignment = Alignment(horizontal='left', vertical='center')
  
   last_col = dcf_worksheet.max_column
  
   styleModule.unbold_category_names(dcf_worksheet)
   styleModule.set_data_cells_background(dcf_worksheet, 9, 20, 1, 2, "F8F7F7")
   dcf.format_rows(dcf_worksheet)
   styleModule.set_data_cells_background(dcf_worksheet, 2, 5, 1, last_col, "F8F7F7")
  
  
   title_row = 8
   start_column = 1  # Column A
   end_column = 2    # Column B


   dcf_worksheet.merge_cells(start_row=title_row, start_column=start_column, end_row=title_row, end_column=end_column)


   # Set the value for the merged cell
   title_cell = dcf_worksheet.cell(row=title_row, column=start_column)
   title_cell.value = "Implied Share Price Calculation"


   title_font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_alignment = Alignment(horizontal='center', vertical='center')
   # Apply the style to the entire merged cell
   for col in range(start_column, end_column + 1):
       cell = dcf_worksheet.cell(row=title_row, column=col)
       cell.font = title_font
       cell.fill = title_fill
       cell.alignment = title_alignment


   title_row = 1
   title_cell = dcf_worksheet.cell(row=title_row, column=1)
   title_cell.value = "Unlevered Free Cash Flow (mm)"
   title_cell.font = Font(color=Color("FFFFFF"), bold=True, size=14)  # White text, bold, and larger font size
   title_cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")  # Light blue background
   title_cell.alignment = Alignment(horizontal='center', vertical='center')
   for col in range(1, last_col + 1):
       cell = dcf_worksheet.cell(row=title_row, column=col)
       cell.fill = PatternFill(start_color="1A759C", end_color="1A759C", fill_type="solid")

  


print("Wrote file")



